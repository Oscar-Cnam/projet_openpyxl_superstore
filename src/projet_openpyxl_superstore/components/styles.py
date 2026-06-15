from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def apply_main_title(ws_visualisations):
    """ Définit le titre du reporting """
    ws_visualisations["K1"] = "Performances économique"
    ws_visualisations["K1"].font = Font(bold=True, size=28, color="185EB8")


def design_numerical_kpis(ws_visualisations, cell, nom):
    """ Définit un style aux titres des indictaurs numériques """
    ws_visualisations[cell] = nom
    ws_visualisations[cell].font = Font(bold=True, size=16)


def create_title_abc(ws_visualisations, cell_ref, letter, color):
    """ Désign des cellule titre (A, B ou C) """
    ws_visualisations[cell_ref] = letter
    ws_visualisations[cell_ref].font = Font(bold=True, size=14)
    ws_visualisations[cell_ref].fill = PatternFill(start_color=color, fill_type="solid")
    ws_visualisations[cell_ref].alignment = Alignment(horizontal="center", vertical="center")


def design_filter_section(ws_visualisations):
    """ Définit un style (couleur de fond et bordure) pour la partie de filtres """
    ws_visualisations.column_dimensions['A'].width = 16.5
    ws_visualisations.column_dimensions['B'].width = 17.56

    ws_visualisations.merge_cells('A4:B4')
    ws_visualisations['A4'] = "Filtres"
    ws_visualisations['A4'].font = Font(bold=True, size=15)
    ws_visualisations['A4'].alignment = Alignment(horizontal="center", vertical="center")

    fond_gris = PatternFill(start_color="F2F2F2", fill_type="solid")
    bordure_fine = Border(
        left=Side(border_style="thin", color="A6A6A6"),
        right=Side(border_style="thin", color="A6A6A6"),
        top=Side(border_style="thin", color="A6A6A6"),
        bottom=Side(border_style="thin", color="A6A6A6")
    )

    for num_cell in range(4, 24):
        ws_visualisations[f'A{num_cell}'].fill = fond_gris
        ws_visualisations[f'B{num_cell}'].fill = fond_gris

    ws_visualisations['A4'].border = bordure_fine
    ws_visualisations['B4'].border = bordure_fine


def apply_all_styles(ws_visualisations):
    """ Fonction orchestratrice pour appliquer les styles """
    apply_main_title(ws_visualisations)
    
    # KPIs déplacés
    design_numerical_kpis(ws_visualisations, "C3", "Chiffre d'affaires")
    design_numerical_kpis(ws_visualisations, "E3", "Profits")
    design_numerical_kpis(ws_visualisations, "G3", "Volume des ventes")
    design_numerical_kpis(ws_visualisations, "J3", "Temps de livraison moyen (en jours)")
    
    # Méthode ABC
    create_title_abc(ws_visualisations, "K6", "A", "56E871") 
    create_title_abc(ws_visualisations, "N6", "B", "E8DC56") 
    create_title_abc(ws_visualisations, "P6", "C", "E87B56") 
    
    # Filtres
    design_filter_section(ws_visualisations)